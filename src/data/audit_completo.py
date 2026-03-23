"""
=============================================================
AUDITORIA COMPLETA DE DATOS - SUENO DORADO ISD
Data Engineer + Data Analyst Senior - Nivel Auditor
=============================================================
Script: audit_completo.py
Fecha: 2026-03-23
Objetivo: Validacion exhaustiva de extraccion, integridad,
          dropdowns, transformaciones y cross-checks.
=============================================================
"""

import sys
import json
from pathlib import Path
from datetime import datetime
from collections import OrderedDict

# Setup paths
ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# --- CONFIG ---------------------------------------------------
EXCEL_PATH = ROOT / "data" / "clean" / "data-main" / "SALDO - FLUJO DE PAGOS 2026.xlsx"
SHEET_NAME = "DATA ORIGINAL"
HEADER_ROW = 3  # 1-indexed, row 3 = header (skiprows=2 in pandas)
CLEAN_CSV = ROOT / "data" / "clean" / "datos_limpios.csv"
CLEAN_XLSX = ROOT / "data" / "clean" / "datos_limpios.xlsx"
VALID_VALUES_JSON = ROOT / "config" / "valid_values.json"
OUTPUT_DIR = ROOT / "reports" / "audit"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Load valid values catalog
with open(VALID_VALUES_JSON, 'r', encoding='utf-8') as f:
    CATALOG = json.load(f)

# --- RESULTS COLLECTOR ----------------------------------------
class AuditCollector:
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.info = []
        self.sections = OrderedDict()

    def error(self, section, msg, details=None):
        entry = {"level": "ERROR", "section": section, "message": msg, "details": details}
        self.errors.append(entry)
        print(f"  [ERROR] | {msg}")

    def warning(self, section, msg, details=None):
        entry = {"level": "WARNING", "section": section, "message": msg, "details": details}
        self.warnings.append(entry)
        print(f"  [WARN]  | {msg}")

    def ok(self, section, msg, details=None):
        entry = {"level": "OK", "section": section, "message": msg, "details": details}
        self.info.append(entry)
        print(f"  [OK]    | {msg}")

    def status(self):
        if self.errors:
            return "ERROR"
        if self.warnings:
            return "WARNING"
        return "OK"

audit = AuditCollector()

# ===============================================================
# TASK 1: VALIDACION DE ESTRUCTURA
# ===============================================================
print("\n" + "="*70)
print("  TASK 1: VALIDACION DE ESTRUCTURA")
print("="*70)

# 1a. Read with openpyxl for raw access
wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
ws = wb[SHEET_NAME]

# Sheet dimensions
opx_max_row = ws.max_row
opx_max_col = ws.max_column
print(f"\n  Dimensiones del sheet: {opx_max_row} filas x {opx_max_col} columnas")

# All sheet names
all_sheets = wb.sheetnames
print(f"  Hojas disponibles: {all_sheets}")
audit.ok("ESTRUCTURA", f"Archivo tiene {len(all_sheets)} hojas: {all_sheets}")

# Get header row raw (row 3 = HEADER_ROW)
raw_headers = []
for col_idx in range(1, opx_max_col + 1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx)
    raw_headers.append({
        "col_num": col_idx,
        "col_letter": openpyxl.utils.get_column_letter(col_idx),
        "value": str(cell.value).strip() if cell.value is not None else None,
        "hidden": ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].hidden
    })

print(f"\n  Columnas encontradas en fila {HEADER_ROW}:")
for h in raw_headers:
    status = " [OCULTA]" if h["hidden"] else ""
    status += " [VACIA]" if h["value"] is None else ""
    print(f"    {h['col_letter']}{h['col_num']:>3}: {h['value']}{status}")

# Detect hidden columns
hidden_cols = [h for h in raw_headers if h["hidden"]]
if hidden_cols:
    audit.warning("ESTRUCTURA", f"{len(hidden_cols)} columna(s) oculta(s) detectada(s): {[h['value'] for h in hidden_cols]}")
else:
    audit.ok("ESTRUCTURA", "Sin columnas ocultas")

# Detect empty header columns
empty_header_cols = [h for h in raw_headers if h["value"] is None]
if empty_header_cols:
    audit.warning("ESTRUCTURA", f"{len(empty_header_cols)} columna(s) sin encabezado: letras {[h['col_letter'] for h in empty_header_cols]}")
else:
    audit.ok("ESTRUCTURA", "Todas las columnas tienen encabezado")

# 1b. Read with pandas for data analysis
df_raw = pd.read_excel(str(EXCEL_PATH), sheet_name=SHEET_NAME, header=2)
df_raw.columns = [str(c).strip() for c in df_raw.columns]

pandas_columns = list(df_raw.columns)
opx_column_names = [h["value"] for h in raw_headers if h["value"] is not None]

print(f"\n  Pandas detecto {len(pandas_columns)} columnas, {len(df_raw)} filas")
print(f"  Openpyxl detecto {len(opx_column_names)} encabezados no-nulos")

# Compare pandas vs openpyxl columns
missing_in_pandas = set(opx_column_names) - set(pandas_columns)
extra_in_pandas = set(pandas_columns) - set(opx_column_names)

if missing_in_pandas:
    audit.error("ESTRUCTURA", f"Columnas en Excel NO presentes en pandas: {missing_in_pandas}")
else:
    audit.ok("ESTRUCTURA", "Todas las columnas del Excel fueron extraidas por pandas")

if extra_in_pandas:
    audit.warning("ESTRUCTURA", f"Columnas extra en pandas (posible rename): {extra_in_pandas}")

# 1c. Data types analysis
print(f"\n  Tipos de datos detectados:")
type_table = []
for col in df_raw.columns:
    detected_type = str(df_raw[col].dtype)
    non_null = df_raw[col].notna().sum()
    null_count = df_raw[col].isna().sum()
    unique_count = df_raw[col].nunique()

    # Detect entirely empty columns
    if non_null == 0:
        estado = "[WARN] VACIA"
        audit.warning("ESTRUCTURA", f"Columna '{col}' esta completamente vacia")
    else:
        estado = "OK"

    type_table.append({
        "columna": col,
        "tipo_detectado": detected_type,
        "no_nulos": int(non_null),
        "nulos": int(null_count),
        "valores_unicos": int(unique_count),
        "estado": estado
    })
    print(f"    {col:30s} | {detected_type:12s} | nulos={null_count:>4} | unicos={unique_count:>4} | {estado}")

# Expected types vs actual
expected_types = {
    "NUMERO UNICO": "numeric/string",
    "IMPORTE": "float64",
    "DOLARES": "float64",
    "DIAS VENCIDOS": "float64/int64",
    "Fecha de Vencimiento": "datetime64",
    "MONEDA": "object",
    "BANCO": "object",
    "GIRADOR": "object",
    "ACEPTANTE": "object",
    "PRODUCTO": "object",
    "CONDICION DEUDA": "object",
}

print(f"\n  Verificacion de tipos esperados:")
for col, expected in expected_types.items():
    if col in df_raw.columns:
        actual = str(df_raw[col].dtype)
        match = "[OK]" if any(e in actual for e in expected.split("/")) else "[WARN]"
        print(f"    {col:30s} | esperado: {expected:15s} | actual: {actual:15s} | {match}")
        if match == "":
            audit.warning("ESTRUCTURA", f"Tipo inesperado en '{col}': esperado={expected}, actual={actual}")

# ===============================================================
# TASK 2: AUDITORIA DE EXTRACCION
# ===============================================================
print("\n" + "="*70)
print("  TASK 2: AUDITORIA DE EXTRACCION")
print("="*70)

# 2a. Row counts
# openpyxl: count data rows (starting after header)
opx_data_rows = 0
for row_idx in range(HEADER_ROW + 1, opx_max_row + 1):
    # Check if at least one cell has a value
    has_data = False
    for col_idx in range(1, min(opx_max_col + 1, 6)):  # Check first 5 cols
        if ws.cell(row=row_idx, column=col_idx).value is not None:
            has_data = True
            break
    if has_data:
        opx_data_rows += 1

pandas_rows = len(df_raw)
print(f"\n  Filas de datos (openpyxl scan): {opx_data_rows}")
print(f"  Filas de datos (pandas):       {pandas_rows}")
print(f"  Diferencia:                    {abs(opx_data_rows - pandas_rows)}")

if opx_data_rows != pandas_rows:
    audit.warning("EXTRACCION", 
        f"Diferencia en conteo: openpyxl={opx_data_rows}, pandas={pandas_rows} (dif={abs(opx_data_rows-pandas_rows)})",
        {"openpyxl": opx_data_rows, "pandas": pandas_rows})
else:
    audit.ok("EXTRACCION", f"Conteo coincide: {pandas_rows} filas")

# 2b. Check for duplicate rows
full_duplicates = df_raw.duplicated().sum()
if full_duplicates > 0:
    audit.warning("EXTRACCION", f"{full_duplicates} filas completamente duplicadas detectadas")
else:
    audit.ok("EXTRACCION", "Sin filas duplicadas completas")

# Check NUMERO UNICO duplicates
if "NUMERO UNICO" in df_raw.columns:
    id_dupes = df_raw["NUMERO UNICO"].dropna().duplicated().sum()
    if id_dupes > 0:
        dupe_vals = df_raw[df_raw["NUMERO UNICO"].duplicated(keep=False)]["NUMERO UNICO"].unique()
        audit.warning("EXTRACCION", 
            f"{id_dupes} IDs duplicados en NUMERO UNICO",
            {"valores_duplicados": [str(v) for v in dupe_vals[:20]]})
    else:
        audit.ok("EXTRACCION", "Sin IDs duplicados en NUMERO UNICO")

# 2c. Truncation check
print(f"\n  Verificacion de truncamiento:")
for col in df_raw.columns:
    if df_raw[col].dtype == "object":
        max_len = df_raw[col].dropna().astype(str).str.len().max()
        if pd.notna(max_len) and max_len > 200:
            audit.warning("EXTRACCION", f"Posible truncamiento en '{col}': max_len={max_len}")
        elif pd.notna(max_len):
            print(f"    {col:30s} | max_chars={int(max_len):>5} [OK]")

# 2d. Compare with cleaned data if exists
if CLEAN_CSV.exists():
    df_clean = pd.read_csv(CLEAN_CSV)
    print(f"\n  Datos limpios encontrados: {CLEAN_CSV.name}")
    print(f"    Filas raw:     {len(df_raw)}")
    print(f"    Filas limpias: {len(df_clean)}")
    filas_perdidas = len(df_raw) - len(df_clean)
    if filas_perdidas > 0:
        audit.warning("EXTRACCION", 
            f"{filas_perdidas} filas perdidas durante limpieza (raw={len(df_raw)} -> clean={len(df_clean)})")
    elif filas_perdidas < 0:
        audit.error("EXTRACCION",
            f"Datos limpios tienen MAS filas que raw ({len(df_clean)} > {len(df_raw)})")
else:
    df_clean = None
    audit.warning("EXTRACCION", f"No se encontro {CLEAN_CSV.name} para comparar")

# ===============================================================
# TASK 3: VALIDACION DE DROPDOWNS (CRITICO)
# ===============================================================
print("\n" + "="*70)
print("  TASK 3: VALIDACION DE DROPDOWNS (CRITICO)")
print("="*70)

# 3a. Extract Excel data validations with openpyxl
data_validations = ws.data_validations.dataValidation if ws.data_validations else []
print(f"\n  Data Validations encontradas en Excel: {len(data_validations)}")

dropdown_analysis = []

for dv in data_validations:
    dv_type = dv.type
    dv_formula = dv.formula1
    dv_cells = str(dv.sqref)
    allow_blank = dv.allow_blank
    
    # Parse the formula to get allowed values
    allowed_values = []
    if dv_formula:
        # Direct list (e.g., "val1,val2,val3")
        if dv_type == "list":
            # Check if it's a range reference or a direct list
            if "!" in str(dv_formula) or "$" in str(dv_formula):
                # Range reference - try to resolve
                allowed_values = [f"REFERENCIA: {dv_formula}"]
            else:
                # Direct comma-separated values
                clean_formula = str(dv_formula).strip('"').strip("'")
                allowed_values = [v.strip() for v in clean_formula.split(",")]
    
    print(f"\n   Validacion: tipo={dv_type}, celdas={dv_cells}")
    print(f"     Formula: {dv_formula}")
    print(f"     Permite vacio: {allow_blank}")
    print(f"     Valores permitidos: {allowed_values}")
    
    # Determine which column this applies to
    # Parse sqref to get column
    try:
        from openpyxl.utils import range_boundaries
        for cell_range in str(dv_cells).split():
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            col_letter = openpyxl.utils.get_column_letter(min_col)
            # Find the header for this column
            header_cell = ws.cell(row=HEADER_ROW, column=min_col)
            col_name = str(header_cell.value).strip() if header_cell.value else f"Col_{col_letter}"
            
            print(f"     -> Columna: {col_name} ({col_letter})")
            
            dropdown_analysis.append({
                "columna": col_name,
                "col_letter": col_letter,
                "tipo_validacion": dv_type,
                "formula": str(dv_formula),
                "valores_permitidos": allowed_values,
                "permite_vacio": allow_blank,
                "rango_celdas": cell_range
            })
    except Exception as e:
        print(f"      Error parseando rango: {e}")

# 3b. Compare catalog values vs actual data
print(f"\n  --- Comparacion: Catalogo vs Datos Reales ---")

dropdown_report = []
for col_name, catalog_values in CATALOG.items():
    if col_name in df_raw.columns:
        actual_values = sorted(df_raw[col_name].dropna().unique().tolist())
        actual_set = set(str(v).strip() for v in actual_values)
        catalog_set = set(str(v).strip() for v in catalog_values)
        
        fuera_catalogo = actual_set - catalog_set
        faltantes = catalog_set - actual_set
        
        # Check for near-matches (case, spaces, typos)
        inconsistencias = []
        for actual in actual_set:
            for cat in catalog_set:
                if actual != cat and (
                    actual.upper() == cat.upper() or  # Case difference
                    actual.strip() == cat.strip() or  # Space difference
                    actual.replace(" ", "") == cat.replace(" ", "")  # Extra spaces
                ):
                    inconsistencias.append(f"'{actual}'  '{cat}' (posible typo/case)")
        
        entry = {
            "columna": col_name,
            "valores_permitidos": sorted(catalog_values),
            "valores_encontrados": sorted([str(v) for v in actual_values]),
            "fuera_catalogo": sorted(fuera_catalogo),
            "faltantes_en_data": sorted(faltantes),
            "inconsistencias": inconsistencias,
            "total_en_data": len(actual_values),
            "total_catalogo": len(catalog_values)
        }
        dropdown_report.append(entry)
        
        print(f"\n  [ANALISIS] {col_name}")
        print(f"     Catalogo ({len(catalog_values)}): {sorted(catalog_values)}")
        print(f"     En data  ({len(actual_values)}): {sorted([str(v) for v in actual_values])}")
        
        if fuera_catalogo:
            audit.warning("DROPDOWN", 
                f"{col_name}: {len(fuera_catalogo)} valor(es) FUERA del catalogo: {sorted(fuera_catalogo)}")
        else:
            audit.ok("DROPDOWN", f"{col_name}: Todos los valores estan en el catalogo")
        
        if faltantes:
            audit.warning("DROPDOWN", 
                f"{col_name}: {len(faltantes)} valor(es) del catalogo NO aparecen en data: {sorted(faltantes)}")
        
        if inconsistencias:
            audit.warning("DROPDOWN", 
                f"{col_name}: Inconsistencias detectadas: {inconsistencias}")
    else:
        audit.error("DROPDOWN", f"Columna '{col_name}' del catalogo NO existe en el dataset")

# ===============================================================
# TASK 4: INTEGRIDAD DE DATOS
# ===============================================================
print("\n" + "="*70)
print("  TASK 4: INTEGRIDAD DE DATOS")
print("="*70)

# 4a. Nulls and empty strings
print(f"\n  --- Analisis de Nulos y Vacios ---")
for col in df_raw.columns:
    nulls = int(df_raw[col].isna().sum())
    empty_str = 0
    if df_raw[col].dtype == "object":
        empty_str = int((df_raw[col].fillna("") == "").sum())
    
    total_missing = nulls + empty_str
    pct = total_missing / len(df_raw) * 100
    
    if pct > 50:
        audit.warning("INTEGRIDAD", f"'{col}': {pct:.1f}% vacio/nulo ({nulls} nulls + {empty_str} vacios)")
    elif pct > 0:
        print(f"    {col:30s} | nulls={nulls:>4} | vacios={empty_str:>4} | {pct:.1f}%")

# 4b. Numeric validity
print(f"\n  --- Validacion Numerica ---")
numeric_cols = ["IMPORTE", "DOLARES", "DIAS VENCIDOS"]
for col in numeric_cols:
    if col not in df_raw.columns:
        continue
    vals = pd.to_numeric(df_raw[col], errors="coerce")
    non_numeric = df_raw[col].notna().sum() - vals.notna().sum()
    negatives = (vals < 0).sum()
    zeros = (vals == 0).sum()
    
    print(f"    {col:20s} | no-numericos={int(non_numeric):>3} | negativos={int(negatives):>3} | ceros={int(zeros):>3}")
    
    if non_numeric > 0:
        # Find specific non-numeric values
        mask = df_raw[col].notna() & vals.isna()
        bad_vals = df_raw.loc[mask, col].unique()[:10]
        audit.error("INTEGRIDAD", 
            f"'{col}': {int(non_numeric)} valores no numericos: {list(bad_vals)}")
    
    if negatives > 0 and col == "IMPORTE":
        audit.error("INTEGRIDAD", f"'{col}': {int(negatives)} valores negativos (regla de negocio)")
    
    if negatives > 0 and col == "DOLARES":
        audit.warning("INTEGRIDAD", f"'{col}': {int(negatives)} valores negativos")

# 4c. Date validation
print(f"\n  --- Validacion de Fechas ---")
if "Fecha de Vencimiento" in df_raw.columns:
    fechas = pd.to_datetime(df_raw["Fecha de Vencimiento"], errors="coerce")
    fechas_invalidas = fechas.isna().sum() - df_raw["Fecha de Vencimiento"].isna().sum()
    
    if fechas.notna().any():
        fecha_min = fechas.min()
        fecha_max = fechas.max()
        print(f"    Rango: {fecha_min} a {fecha_max}")
        
        # Suspicious dates
        muy_antiguas = (fechas < pd.Timestamp("2020-01-01")).sum()
        muy_futuras = (fechas > pd.Timestamp("2030-12-31")).sum()
        
        if muy_antiguas > 0:
            audit.warning("INTEGRIDAD", f"Fechas: {int(muy_antiguas)} fecha(s) antes de 2020")
        if muy_futuras > 0:
            audit.warning("INTEGRIDAD", f"Fechas: {int(muy_futuras)} fecha(s) despues de 2030")
        if int(fechas_invalidas) > 0:
            audit.error("INTEGRIDAD", f"Fechas: {int(fechas_invalidas)} fecha(s) no parseables")
        else:
            audit.ok("INTEGRIDAD", f"Todas las fechas son validas ({fecha_min.date()} a {fecha_max.date()})")

# 4d. Outlier detection (IQR method)
print(f"\n  --- Deteccion de Outliers ---")
if "IMPORTE" in df_raw.columns:
    vals = pd.to_numeric(df_raw["IMPORTE"], errors="coerce").dropna()
    if len(vals) >= 4:
        Q1 = vals.quantile(0.25)
        Q3 = vals.quantile(0.75)
        IQR = Q3 - Q1
        lower = Q1 - 1.5 * IQR
        upper = Q3 + 1.5 * IQR
        outliers_low = vals[vals < lower]
        outliers_high = vals[vals > upper]
        
        print(f"    IMPORTE: Q1={Q1:,.2f}, Q3={Q3:,.2f}, IQR={IQR:,.2f}")
        print(f"    Limites: [{lower:,.2f}, {upper:,.2f}]")
        print(f"    Outliers bajos: {len(outliers_low)}, altos: {len(outliers_high)}")
        
        if len(outliers_high) > 0:
            top_outliers = outliers_high.nlargest(5).tolist()
            audit.warning("INTEGRIDAD", 
                f"IMPORTE: {len(outliers_high)} outliers superiores. Top 5: {[f'S/.{v:,.2f}' for v in top_outliers]}")
        
        # Strict 3*IQR
        upper_strict = Q3 + 3 * IQR
        extreme = vals[vals > upper_strict]
        if len(extreme) > 0:
            audit.warning("INTEGRIDAD", f"IMPORTE: {len(extreme)} outliers EXTREMOS (>3xIQR = S/.{upper_strict:,.2f})")

# ===============================================================
# TASK 5: VALIDACION DE TRANSFORMACIONES
# ===============================================================
print("\n" + "="*70)
print("  TASK 5: VALIDACION DE TRANSFORMACIONES")
print("="*70)

# Filter like app.py does
df_app = df_raw.copy()
if "CONDICION DEUDA" in df_app.columns:
    df_app = df_app[df_app["CONDICION DEUDA"] == "PENDIENTE DE PAGO"]
    print(f"\n  Filtro 'PENDIENTE DE PAGO': {len(df_app)} filas (de {len(df_raw)})")

df_app["IMPORTE_NUM"] = pd.to_numeric(
    df_app["IMPORTE"].astype(str).str.replace(",", ""), errors="coerce"
)
df_app = df_app[df_app["IMPORTE_NUM"].notna()]
print(f"  Tras filtrar IMPORTE valido: {len(df_app)} filas")

# Recalculate key metrics
if "MONEDA" in df_app.columns:
    TC = 3.45  # Exchange rate from app.py default
    
    soles_calc = df_app[df_app["MONEDA"] == "SOLES"]["IMPORTE_NUM"].sum()
    dolares_calc = df_app[df_app["MONEDA"] == "DOLARES"]["IMPORTE_NUM"].sum()
    total_usd_calc = soles_calc / TC + dolares_calc
    
    vencidos_calc = 0
    if "DIAS VENCIDOS" in df_app.columns:
        dias = pd.to_numeric(df_app["DIAS VENCIDOS"].astype(str).str.replace(",", ""), errors="coerce")
        vencidos_calc = int((dias > 0).sum())
        al_dia_calc = int((dias == 0).sum())
        por_vencer_calc = int((dias < 0).sum())
        riesgo_alto_calc = int((dias > 90).sum())
    
    print(f"\n  --- Metricas Recalculadas ---")
    print(f"    Total documentos (filtrados):  {len(df_app)}")
    print(f"    Cartera en Soles:             S/. {soles_calc:,.2f}")
    print(f"    Cartera en Dolares:           $   {dolares_calc:,.2f}")
    print(f"    Total USD (TC={TC}):          $   {total_usd_calc:,.2f}")
    print(f"    Vencidos:                     {vencidos_calc}")
    print(f"    Al dia:                       {al_dia_calc}")
    print(f"    Por vencer:                   {por_vencer_calc}")
    print(f"    Riesgo alto (>90d):           {riesgo_alto_calc}")
    
    audit.ok("TRANSFORMACIONES", 
        f"Metricas recalculadas: {len(df_app)} docs, S/.{soles_calc:,.0f} soles, ${dolares_calc:,.0f} USD")

# Compare with clean data
if df_clean is not None:
    print(f"\n  --- Comparacion Raw vs Limpio ---")
    
    # Column comparison
    raw_cols = set(df_raw.columns)
    if '_limpieza_fecha' in df_clean.columns:
        clean_cols_no_meta = set(c for c in df_clean.columns if not c.startswith('_'))
    else:
        clean_cols_no_meta = set(df_clean.columns)
    
    added_cols = clean_cols_no_meta - raw_cols
    removed_cols = raw_cols - clean_cols_no_meta
    
    if added_cols:
        print(f"    Columnas ANADIDAS en limpieza: {added_cols}")
    if removed_cols:
        print(f"    Columnas ELIMINADAS en limpieza: {removed_cols}")
        audit.warning("TRANSFORMACIONES", f"Columnas eliminadas en limpieza: {removed_cols}")

    # Row count delta
    print(f"    Filas raw: {len(df_raw)} -> limpias: {len(df_clean)} (delta: {len(df_raw)-len(df_clean)})")

# ===============================================================
# TASK 6: CROSS-CHECKS
# ===============================================================
print("\n" + "="*70)
print("  TASK 6: CROSS-CHECKS")
print("="*70)

if "IMPORTE_NUM" in df_app.columns:
    # 6a. Total global vs sum of groups
    total_global = df_app["IMPORTE_NUM"].sum()
    
    # By MONEDA
    if "MONEDA" in df_app.columns:
        total_by_moneda = df_app.groupby("MONEDA")["IMPORTE_NUM"].sum().sum()
        match = abs(total_global - total_by_moneda) < 0.01
        print(f"\n  Total global:       S/. {total_global:,.2f}")
        print(f"  Suma por MONEDA:    S/. {total_by_moneda:,.2f}")
        print(f"  Coincide:           {'[OK]' if match else '[ERROR]'}")
        if not match:
            audit.error("CROSS-CHECK", f"Total global  suma por MONEDA: {total_global:,.2f} vs {total_by_moneda:,.2f}")
        else:
            audit.ok("CROSS-CHECK", "Total global = Suma por MONEDA ")
    
    # By BANCO
    if "BANCO" in df_app.columns:
        total_by_banco = df_app.groupby("BANCO")["IMPORTE_NUM"].sum().sum()
        match = abs(total_global - total_by_banco) < 0.01
        print(f"  Suma por BANCO:     S/. {total_by_banco:,.2f}")
        print(f"  Coincide:           {'[OK]' if match else '[ERROR]'}")
        if not match:
            audit.error("CROSS-CHECK", f"Total global  suma por BANCO")
        else:
            audit.ok("CROSS-CHECK", "Total global = Suma por BANCO ")
    
    # By GIRADOR
    if "GIRADOR" in df_app.columns:
        total_by_girador = df_app.groupby("GIRADOR")["IMPORTE_NUM"].sum().sum()
        match = abs(total_global - total_by_girador) < 0.01
        print(f"  Suma por GIRADOR:   S/. {total_by_girador:,.2f}")
        print(f"  Coincide:           {'[OK]' if match else '[ERROR]'}")
        if not match:
            audit.error("CROSS-CHECK", f"Total global  suma por GIRADOR")
        else:
            audit.ok("CROSS-CHECK", "Total global = Suma por GIRADOR ")
    
    # By PRODUCTO
    if "PRODUCTO" in df_app.columns:
        total_by_producto = df_app.groupby("PRODUCTO")["IMPORTE_NUM"].sum().sum()
        match = abs(total_global - total_by_producto) < 0.01
        print(f"  Suma por PRODUCTO:  S/. {total_by_producto:,.2f}")
        print(f"  Coincide:           {'[OK]' if match else '[ERROR]'}")
        if not match:
            audit.error("CROSS-CHECK", f"Total global  suma por PRODUCTO")
        else:
            audit.ok("CROSS-CHECK", "Total global = Suma por PRODUCTO ")
    
    # By ACEPTANTE
    if "ACEPTANTE" in df_app.columns:
        total_by_aceptante = df_app.groupby("ACEPTANTE")["IMPORTE_NUM"].sum().sum()
        match = abs(total_global - total_by_aceptante) < 0.01
        print(f"  Suma por ACEPTANTE: S/. {total_by_aceptante:,.2f}")
        print(f"  Coincide:           {'[OK]' if match else '[ERROR]'}")
        if not match:
            audit.error("CROSS-CHECK", f"Total global  suma por ACEPTANTE")
        else:
            audit.ok("CROSS-CHECK", "Total global = Suma por ACEPTANTE ")

    # 6b. Count cross-check
    total_docs = len(df_app)
    if "BANCO" in df_app.columns:
        count_by_banco = df_app.groupby("BANCO").size().sum()
        if total_docs != count_by_banco:
            audit.error("CROSS-CHECK", f"Conteo docs  conteo por BANCO: {total_docs} vs {count_by_banco}")
        else:
            audit.ok("CROSS-CHECK", f"Conteo de docs coincide por BANCO: {total_docs}")

    # 6c. DOLARES vs IMPORTE consistency for dolares
    if "DOLARES" in df_app.columns and "MONEDA" in df_app.columns:
        df_dol = df_app[df_app["MONEDA"] == "DOLARES"].copy()
        df_dol["DOLARES_NUM"] = pd.to_numeric(df_dol["DOLARES"].astype(str).str.replace(",", ""), errors="coerce")
        
        importe_dol = df_dol["IMPORTE_NUM"].sum()
        dolares_col = df_dol["DOLARES_NUM"].sum()
        
        print(f"\n  --- Consistencia DOLARES ---")
        print(f"    IMPORTE (moneda=DOLARES): S/. {importe_dol:,.2f}")
        print(f"    Columna DOLARES:          $   {dolares_col:,.2f}")
        
        if abs(importe_dol - dolares_col) > 1:
            audit.warning("CROSS-CHECK", 
                f"IMPORTE vs DOLARES difieren para moneda DOLARES: {importe_dol:,.2f} vs {dolares_col:,.2f}")

# ===============================================================
# TASK 7: REPORTE FINAL DE INCONSISTENCIAS
# ===============================================================
print("\n" + "="*70)
print("  TASK 7: REPORTE DE INCONSISTENCIAS")
print("="*70)

status = audit.status()
indicator = "[ERROR]" if status == "ERROR" else "[WARN]" if status == "WARNING" else "[OK]"
print(f"\n  {indicator} ESTADO GENERAL: {status}")
print(f"\n  Errores:      {len(audit.errors)}")
print(f"  Advertencias: {len(audit.warnings)}")
print(f"  OK:           {len(audit.info)}")

if audit.errors:
    print(f"\n  --- ERRORES CRITICOS ---")
    for i, e in enumerate(audit.errors, 1):
        print(f"    {i}. [{e['section']}] {e['message']}")

if audit.warnings:
    print(f"\n  --- ADVERTENCIAS ---")
    for i, w in enumerate(audit.warnings, 1):
        print(f"    {i}. [{w['section']}] {w['message']}")

# --- SAVE FULL REPORT -----------------------------------------
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
report = {
    "meta": {
        "timestamp": datetime.now().isoformat(),
        "archivo_auditado": str(EXCEL_PATH),
        "hoja": SHEET_NAME,
        "estado_general": status,
        "total_errores": len(audit.errors),
        "total_advertencias": len(audit.warnings),
        "total_ok": len(audit.info),
    },
    "estructura": {
        "columnas_excel": [h["value"] for h in raw_headers],
        "columnas_pandas": pandas_columns,
        "columnas_ocultas": [h["value"] for h in hidden_cols],
        "columnas_vacias": [h["col_letter"] for h in empty_header_cols],
        "tabla_tipos": type_table,
    },
    "extraccion": {
        "filas_openpyxl": opx_data_rows,
        "filas_pandas": pandas_rows,
        "duplicados_completos": int(full_duplicates),
    },
    "dropdowns": {
        "validaciones_excel": [
            {
                "columna": d["columna"],
                "tipo": d["tipo_validacion"],
                "formula": d["formula"],
                "valores_permitidos": d["valores_permitidos"],
            }
            for d in dropdown_analysis
        ],
        "analisis_catalogo": dropdown_report,
    },
    "errores": audit.errors,
    "advertencias": audit.warnings,
    "info": audit.info,
}

report_path = OUTPUT_DIR / f"auditoria_completa_{timestamp}.json"
with open(report_path, "w", encoding="utf-8") as f:
    json.dump(report, f, ensure_ascii=False, indent=2, default=str)

print(f"\n   Reporte guardado: {report_path.name}")
print("\n" + "="*70)
print("  AUDITORIA COMPLETA FINALIZADA")
print("="*70)
